import openpyxl
import pandas as pd
from django.core.exceptions import ValidationError
from django.db import transaction
from django.db.models import Sum, Avg, F, Q, Count
from django.utils import timezone
from django.core.mail import send_mail, EmailMessage
from django.conf import settings
from django.template.loader import render_to_string
from fpdf import FPDF
from datetime import timedelta, date
import logging
import io

from .models import TestCase, UserStory, TestExecution, TestRun, Project

logger = logging.getLogger(__name__)


def validate_excel_headers(worksheet):
    """
    Validate that the Excel file has the required headers.
    Expected headers: Test Case Name, Description, Test Steps, Expected Results, Priority, Is Automated
    """
    required_headers = [
        'Test Case Name',
        'Description', 
        'Test Steps',
        'Expected Results',
        'Priority',
        'Is Automated'
    ]
    
    # Get first row (headers)
    headers = []
    for cell in worksheet[1]:
        if cell.value:
            headers.append(str(cell.value).strip())
    
    missing_headers = []
    for required_header in required_headers:
        if required_header not in headers:
            missing_headers.append(required_header)
    
    if missing_headers:
        return False, f"Missing required headers: {', '.join(missing_headers)}"
    
    return True, headers


def validate_testcase_data(row_data, row_number):
    """
    Validate individual test case data from Excel row.
    Returns tuple (is_valid, errors_list)
    """
    errors = []
    
    # Check required fields
    if not row_data.get('Test Case Name', '').strip():
        errors.append(f"Row {row_number}: Test Case Name is required")
    elif len(row_data.get('Test Case Name', '').strip()) < 5:
        errors.append(f"Row {row_number}: Test Case Name must be at least 5 characters long")
    
    if not row_data.get('Test Steps', '').strip():
        errors.append(f"Row {row_number}: Test Steps are required")
    elif len(row_data.get('Test Steps', '').strip()) < 10:
        errors.append(f"Row {row_number}: Test Steps must be at least 10 characters long")
    
    if not row_data.get('Expected Results', '').strip():
        errors.append(f"Row {row_number}: Expected Results are required")
    elif len(row_data.get('Expected Results', '').strip()) < 10:
        errors.append(f"Row {row_number}: Expected Results must be at least 10 characters long")
    
    # Validate priority
    valid_priorities = ['low', 'medium', 'high', 'critical']
    priority = str(row_data.get('Priority', 'medium')).lower().strip()
    if priority and priority not in valid_priorities:
        errors.append(f"Row {row_number}: Priority must be one of: {', '.join(valid_priorities)}")
    
    # Validate is_automated
    is_automated_value = str(row_data.get('Is Automated', 'False')).lower().strip()
    if is_automated_value not in ['true', 'false', '1', '0', 'yes', 'no']:
        errors.append(f"Row {row_number}: Is Automated must be True/False, Yes/No, or 1/0")
    
    return len(errors) == 0, errors


def parse_excel_file(excel_file):
    """
    Parse Excel file and return list of test case data dictionaries.
    Returns tuple (success, data_list, error_message)
    """
    try:
        # Try to read with openpyxl first
        workbook = openpyxl.load_workbook(excel_file, data_only=True)
        worksheet = workbook.active
        
        # Validate headers
        headers_valid, headers_or_error = validate_excel_headers(worksheet)
        if not headers_valid:
            return False, [], headers_or_error
        
        headers = headers_or_error
        
        # Parse data rows
        data_list = []
        errors = []
        
        for row_idx, row in enumerate(worksheet.iter_rows(min_row=2, values_only=True), start=2):
            if not any(row):  # Skip empty rows
                continue
            
            # Create dictionary from row data
            row_data = {}
            for col_idx, header in enumerate(headers):
                if col_idx < len(row):
                    cell_value = row[col_idx]
                    row_data[header] = str(cell_value) if cell_value is not None else ''
                else:
                    row_data[header] = ''
            
            # Validate row data
            is_valid, validation_errors = validate_testcase_data(row_data, row_idx)
            if is_valid:
                data_list.append(row_data)
            else:
                errors.extend(validation_errors)
        
        if errors:
            return False, [], f"Validation errors found:\n" + "\n".join(errors)
        
        return True, data_list, ""
        
    except Exception as e:
        logger.error(f"Error parsing Excel file: {str(e)}")
        return False, [], f"Error reading Excel file: {str(e)}"


def convert_boolean_value(value):
    """
    Convert various boolean representations to Python boolean.
    """
    if isinstance(value, bool):
        return value
    
    value_str = str(value).lower().strip()
    return value_str in ['true', '1', 'yes', 'y']


def normalize_priority(priority_value):
    """
    Normalize priority value to valid choice.
    """
    priority_str = str(priority_value).lower().strip()
    priority_mapping = {
        'low': 'low',
        'l': 'low',
        'medium': 'medium',
        'med': 'medium',
        'm': 'medium',
        'high': 'high',
        'h': 'high',
        'critical': 'critical',
        'crit': 'critical',
        'c': 'critical'
    }
    
    return priority_mapping.get(priority_str, 'medium')


def create_testcase_from_data(testcase_data, user_story, created_by):
    """
    Create a TestCase instance from parsed Excel data.
    Returns tuple (testcase_instance, success, error_message)
    """
    try:
        # Normalize data
        name = testcase_data.get('Test Case Name', '').strip()
        description = testcase_data.get('Description', '').strip()
        test_steps = testcase_data.get('Test Steps', '').strip()
        expected_results = testcase_data.get('Expected Results', '').strip()
        priority = normalize_priority(testcase_data.get('Priority', 'medium'))
        is_automated = convert_boolean_value(testcase_data.get('Is Automated', 'False'))
        
        # Create TestCase instance
        testcase = TestCase(
            name=name,
            description=description,
            test_steps=test_steps,
            expected_results=expected_results,
            priority=priority,
            is_automated=is_automated,
            user_story=user_story,
            created_by=created_by,
            status='draft',
            execution_status='not_executed'
        )
        
        # Validate the instance
        testcase.full_clean()
        
        return testcase, True, ""
        
    except ValidationError as e:
        error_msg = "; ".join([f"{field}: {', '.join(errors)}" for field, errors in e.message_dict.items()])
        return None, False, f"Validation error: {error_msg}"
    except Exception as e:
        return None, False, f"Error creating test case: {str(e)}"


@transaction.atomic
def import_testcases_from_excel(excel_file, user_story, created_by):
    """
    Main function to import test cases from Excel file.
    Returns dictionary with results: {
        'success': bool,
        'created': int,
        'errors': list,
        'error': str (if success is False)
    }
    """
    result = {
        'success': False,
        'created': 0,
        'errors': [],
        'error': ''
    }
    
    try:
        # Validate user_story exists and user has permission
        if not isinstance(user_story, UserStory):
            result['error'] = "Invalid user story provided"
            return result
        
        # Check if user has permission to create test cases for this user story
        if user_story.epic.project.created_by != created_by:
            result['error'] = "You don't have permission to create test cases for this user story"
            return result
        
        # Parse Excel file
        parse_success, testcase_data_list, parse_error = parse_excel_file(excel_file)
        if not parse_success:
            result['error'] = parse_error
            return result
        
        if not testcase_data_list:
            result['error'] = "No valid test case data found in Excel file"
            return result
        
        # Create test cases
        created_testcases = []
        errors = []
        
        for idx, testcase_data in enumerate(testcase_data_list, start=1):
            testcase, create_success, create_error = create_testcase_from_data(
                testcase_data, user_story, created_by
            )
            
            if create_success:
                created_testcases.append(testcase)
            else:
                errors.append(f"Row {idx + 1}: {create_error}")
        
        # Bulk create test cases if any were successfully prepared
        if created_testcases:
            try:
                TestCase.objects.bulk_create(created_testcases)
                result['created'] = len(created_testcases)
                logger.info(f"Successfully imported {len(created_testcases)} test cases for user {created_by.username}")
            except Exception as e:
                result['error'] = f"Error saving test cases to database: {str(e)}"
                return result
        
        result['success'] = True
        result['errors'] = errors
        
        # Log summary
        if errors:
            logger.warning(f"Import completed with {len(errors)} errors for user {created_by.username}")
        
        return result
        
    except Exception as e:
        logger.error(f"Unexpected error during Excel import: {str(e)}")
        result['error'] = f"Unexpected error during import: {str(e)}"
        return result


def generate_sample_excel_data():
    """
    Generate sample data for Excel template download.
    Returns list of dictionaries representing sample test cases.
    """
    sample_data = [
        {
            'Test Case Name': 'Login with valid credentials',
            'Description': 'Test successful login with valid username and password',
            'Test Steps': '1. Navigate to login page\\n2. Enter valid username\\n3. Enter valid password\\n4. Click login button',
            'Expected Results': 'User should be successfully logged in and redirected to dashboard',
            'Priority': 'high',
            'Is Automated': 'False'
        },
        {
            'Test Case Name': 'Login with invalid credentials',
            'Description': 'Test login failure with invalid credentials',
            'Test Steps': '1. Navigate to login page\\n2. Enter invalid username\\n3. Enter invalid password\\n4. Click login button',
            'Expected Results': 'Error message should be displayed and user should remain on login page',
            'Priority': 'medium',
            'Is Automated': 'True'
        },
        {
            'Test Case Name': 'Password reset functionality',
            'Description': 'Test password reset feature',
            'Test Steps': '1. Click forgot password link\\n2. Enter email address\\n3. Click submit\\n4. Check email for reset link',
            'Expected Results': 'Password reset email should be sent with valid reset link',
            'Priority': 'low',
            'Is Automated': 'False'
        }
    ]
    
    return sample_data


def export_testcases_to_excel(testcases, filename=None):
    """
    Export test cases to Excel format.
    Returns tuple (success, file_path_or_error_message)
    """
    try:
        # Create workbook and worksheet
        workbook = openpyxl.Workbook()
        worksheet = workbook.active
        worksheet.title = "Test Cases"
        
        # Define headers
        headers = [
            'Test Case Name', 'Description', 'Test Steps', 'Expected Results',
            'Priority', 'Status', 'Execution Status', 'Is Automated',
            'User Story', 'Epic', 'Project', 'Created Date', 'Updated Date'
        ]
        
        # Write headers
        for col_idx, header in enumerate(headers, start=1):
            worksheet.cell(row=1, column=col_idx, value=header)
        
        # Write test case data
        for row_idx, testcase in enumerate(testcases, start=2):
            worksheet.cell(row=row_idx, column=1, value=testcase.name)
            worksheet.cell(row=row_idx, column=2, value=testcase.description)
            worksheet.cell(row=row_idx, column=3, value=testcase.test_steps)
            worksheet.cell(row=row_idx, column=4, value=testcase.expected_results)
            worksheet.cell(row=row_idx, column=5, value=testcase.get_priority_display())
            worksheet.cell(row=row_idx, column=6, value=testcase.get_status_display())
            worksheet.cell(row=row_idx, column=7, value=testcase.get_execution_status_display())
            worksheet.cell(row=row_idx, column=8, value='Yes' if testcase.is_automated else 'No')
            worksheet.cell(row=row_idx, column=9, value=testcase.user_story.name)
            worksheet.cell(row=row_idx, column=10, value=testcase.user_story.epic.name)
            worksheet.cell(row=row_idx, column=11, value=testcase.user_story.epic.project.name)
            worksheet.cell(row=row_idx, column=12, value=testcase.created_date.strftime('%Y-%m-%d %H:%M:%S'))
            worksheet.cell(row=row_idx, column=13, value=testcase.updated_date.strftime('%Y-%m-%d %H:%M:%S'))
        
        # Auto-adjust column widths
        for column in worksheet.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            worksheet.column_dimensions[column_letter].width = adjusted_width
        
        # Save to BytesIO if no filename provided (for HTTP response)
        if not filename:
            excel_buffer = io.BytesIO()
            workbook.save(excel_buffer)
            excel_buffer.seek(0)
            return True, excel_buffer
        else:
            workbook.save(filename)
            return True, filename
            
    except Exception as e:
        logger.error(f"Error exporting test cases to Excel: {str(e)}")
        return False, f"Error exporting to Excel: {str(e)}"


def generate_execution_excel_report(executions):
    """
    Generates an Excel report for a given queryset of TestExecution objects.
    Returns a BytesIO object containing the Excel file.
    """
    try:
        workbook = openpyxl.Workbook()
        worksheet = workbook.active
        worksheet.title = "Execution Report"

        headers = [
            "Test Run", "Test Case", "Project", "Epic", "User Story", "Status",
            "Executor", "Execution Date", "Duration (min)", "Comments", "Notes"
        ]
        worksheet.append(headers)

        for execution in executions:
            row_data = [
                execution.test_run.name,
                execution.testcase.name,
                execution.testcase.user_story.epic.project.name,
                execution.testcase.user_story.epic.name,
                execution.testcase.user_story.name,
                execution.get_status_display(),
                execution.executor.get_full_name() if execution.executor else "N/A",
                execution.execution_date.strftime("%Y-%m-%d %H:%M:%S") if execution.execution_date else "N/A",
                execution.execution_time_minutes if execution.execution_time_minutes is not None else "N/A",
                execution.comments,
                execution.notes,
            ]
            worksheet.append(row_data)
        
        # Auto-adjust column widths
        for column in worksheet.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 70)
            worksheet.column_dimensions[column_letter].width = adjusted_width

        excel_buffer = io.BytesIO()
        workbook.save(excel_buffer)
        excel_buffer.seek(0)
        logger.info("Successfully generated execution Excel report.")
        return excel_buffer

    except Exception as e:
        logger.error(f"Error generating execution Excel report: {e}")
        return None

def calculate_execution_metrics(executions_queryset):
    """
    Calculates execution metrics for a given queryset of TestExecution objects.
    Returns a dictionary with various statistics.
    """
    metrics = {
        'total': executions_queryset.count(),
        'passed': 0,
        'failed': 0,
        'skipped': 0,
        'blocked': 0,
        'not_executed': 0,
        'in_progress': 0,
        'pass_rate': 0.0,
        'avg_execution_time': 0.0,
        'total_execution_time': 0,
        'unique_executors': 0,
        'projects_breakdown': [],
        'assignee_performance': [],
        'trends': {'labels': [], 'total': [], 'passed': [], 'failed': []}
    }

    if metrics['total'] == 0:
        return metrics

    try:
        metrics['passed'] = executions_queryset.filter(status='passed').count()
        metrics['failed'] = executions_queryset.filter(status='failed').count()
        metrics['skipped'] = executions_queryset.filter(status='skipped').count()
        metrics['blocked'] = executions_queryset.filter(status='blocked').count()
        metrics['not_executed'] = executions_queryset.filter(status='not_executed').count()
        metrics['in_progress'] = executions_queryset.filter(status='in_progress').count()

        executed_count = metrics['passed'] + metrics['failed'] + metrics['skipped'] + metrics['blocked']
        if executed_count > 0:
            metrics['pass_rate'] = (metrics['passed'] / executed_count) * 100

        avg_time_agg = executions_queryset.filter(
            execution_time_minutes__isnull=False
        ).aggregate(avg_time=Avg('execution_time_minutes'))
        metrics['avg_execution_time'] = avg_time_agg['avg_time'] if avg_time_agg['avg_time'] is not None else 0.0

        total_time_agg = executions_queryset.filter(
            execution_time_minutes__isnull=False
        ).aggregate(total_time=Sum('execution_time_minutes'))
        metrics['total_execution_time'] = total_time_agg['total_time'] if total_time_agg['total_time'] is not None else 0

        metrics['unique_executors'] = executions_queryset.filter(executor__isnull=False).values('executor').distinct().count()

        # Project-wise breakdown
        project_stats = Project.objects.filter(
            epics__user_stories__test_cases__test_executions__in=executions_queryset
        ).annotate(
            total_executions=Count('epics__user_stories__test_cases__test_executions', distinct=True),
            passed_executions=Count('epics__user_stories__test_cases__test_executions', filter=Q(epics__user_stories__test_cases__test_executions__status='passed'), distinct=True),
            failed_executions=Count('epics__user_stories__test_cases__test_executions', filter=Q(epics__user_stories__test_cases__test_executions__status='failed'), distinct=True),
            avg_time=Avg('epics__user_stories__test_cases__test_executions__execution_time_minutes')
        ).order_by('-total_executions')

        for project in project_stats:
            pass_rate = (project.passed_executions / (project.passed_executions + project.failed_executions) * 100) if (project.passed_executions + project.failed_executions) > 0 else 0
            metrics['projects_breakdown'].append({
                'id': project.id,
                'name': project.name,
                'total_executions': project.total_executions,
                'passed_executions': project.passed_executions,
                'failed_executions': project.failed_executions,
                'pass_rate': pass_rate,
                'avg_execution_time': project.avg_time if project.avg_time is not None else 0.0
            })
        
        # Assignee performance
        assignee_stats = executions_queryset.filter(executor__isnull=False).values(
            'executor__first_name', 'executor__last_name', 'executor__username'
        ).annotate(
            executions_count=Count('id'),
            passed_count=Count('id', filter=Q(status='passed')),
            failed_count=Count('id', filter=Q(status='failed')),
            avg_time=Avg('execution_time_minutes')
        ).order_by('-executions_count')

        for assignee in assignee_stats:
            full_name = f"{assignee['executor__first_name']} {assignee['executor__last_name']}".strip()
            name = full_name if full_name else assignee['executor__username']
            assignee_executed = assignee['passed_count'] + assignee['failed_count']
            pass_rate = (assignee['passed_count'] / assignee_executed * 100) if assignee_executed > 0 else 0
            metrics['assignee_performance'].append({
                'name': name,
                'executions_count': assignee['executions_count'],
                'passed_count': assignee['passed_count'],
                'failed_count': assignee['failed_count'],
                'pass_rate': pass_rate,
                'avg_time': assignee['avg_time'] if assignee['avg_time'] is not None else 0.0
            })

        # Execution Trends (last 30 days)
        today = timezone.localdate()
        for i in range(30):
            current_date = today - timedelta(days=29 - i)
            metrics['trends']['labels'].append(current_date.strftime('%Y-%m-%d'))
            daily_executions = executions_queryset.filter(
                execution_date__date=current_date
            )
            metrics['trends']['total'].append(daily_executions.count())
            metrics['trends']['passed'].append(daily_executions.filter(status='passed').count())
            metrics['trends']['failed'].append(daily_executions.filter(status='failed').count())

        logger.info("Successfully calculated execution metrics.")
        return metrics

    except Exception as e:
        logger.error(f"Error calculating execution metrics: {e}")
        return metrics # Return partial metrics in case of error


def send_execution_notification(subject, message, recipient_list, from_email=None, html_message=None):
    """
    Sends an email notification for test execution updates.
    """
    try:
        if not from_email:
            from_email = settings.DEFAULT_FROM_EMAIL
        
        send_mail(
            subject,
            message,
            from_email,
            recipient_list,
            html_message=html_message,
            fail_silently=False,
        )
        logger.info(f"Execution notification sent to {', '.join(recipient_list)}")
        return True
    except Exception as e:
        logger.error(f"Failed to send execution notification: {e}")
        return False

class PDFReport(FPDF):
    def header(self):
        self.set_font('Arial', 'B', 15)
        self.cell(0, 10, 'Test Execution Report', 0, 1, 'C')
        self.ln(10)

    def footer(self):
        self.set_y(-15)
        self.set_font('Arial', 'I', 8)
        self.cell(0, 10, f'Page {self.page_no()}/{{nb}}', 0, 0, 'C')

    def chapter_title(self, title):
        self.set_font('Arial', 'B', 12)
        self.cell(0, 10, title, 0, 1, 'L')
        self.ln(2)

    def chapter_body(self, body):
        self.set_font('Arial', '', 10)
        self.multi_cell(0, 5, body)
        self.ln()

def create_execution_pdf_report(executions_queryset):
    """
    Generates a PDF report for a given queryset of TestExecution objects.
    Returns a BytesIO object containing the PDF file.
    """
    try:
        pdf = PDFReport()
        pdf.alias_nb_pages()
        pdf.add_page()
        pdf.set_auto_page_break(auto=True, margin=15)

        metrics = calculate_execution_metrics(executions_queryset)

        # Summary Section
        pdf.chapter_title("1. Execution Summary")
        summary_text = (
            f"Total Executions: {metrics['total']}\\n"
            f"Passed: {metrics['passed']}\\n"
            f"Failed: {metrics['failed']}\\n"
            f"Skipped: {metrics['skipped']}\\n"
            f"Blocked: {metrics['blocked']}\\n"
            f"Not Executed: {metrics['not_executed']}\\n"
            f"In Progress: {metrics['in_progress']}\\n"
            f"Pass Rate: {metrics['pass_rate']:.2f}%\\n"
            f"Average Execution Time: {metrics['avg_execution_time']:.2f} minutes\\n"
            f"Total Execution Time: {metrics['total_execution_time']} minutes\\n"
        )
        pdf.chapter_body(summary_text)

        # Detailed Executions
        pdf.chapter_title("2. Detailed Execution Results")
        if not executions_queryset.exists():
            pdf.chapter_body("No detailed execution results found for the selected filters.")
        else:
            for i, execution in enumerate(executions_queryset, 1):
                pdf.chapter_title(f"2.{i}. Test Case: {execution.testcase.name} ({execution.test_run.name})")
                details = (
                    f"Project: {execution.testcase.user_story.epic.project.name}\\n"
                    f"Epic: {execution.testcase.user_story.epic.name}\\n"
                    f"User Story: {execution.testcase.user_story.name}\\n"
                    f"Status: {execution.get_status_display()}\\n"
                    f"Executor: {execution.executor.get_full_name() if execution.executor else 'N/A'}\\n"
                    f"Execution Date: {execution.execution_date.strftime('%Y-%m-%d %H:%M:%S') if execution.execution_date else 'N/A'}\\n"
                    f"Duration: {execution.execution_time_minutes} minutes\\n"
                    f"Comments: {execution.comments}\\n"
                    f"Notes: {execution.notes}\\n"
                )
                pdf.chapter_body(details)
        
        pdf_buffer = io.BytesIO()
        pdf.output(pdf_buffer)
        pdf_buffer.seek(0)
        logger.info("Successfully generated PDF execution report.")
        return pdf_buffer
    except Exception as e:
        logger.error(f"Error generating PDF execution report: {e}")
        return None

@transaction.atomic
def bulk_execute_testcases(test_case_ids, status, comments, executor_user, test_run=None):
    """
    Executes multiple test cases with a given status and comments.
    Can create a new TestRun or update an existing one.
    Returns a dictionary with success status and message.
    """
    try:
        if not isinstance(test_case_ids, list):
            raise TypeError("test_case_ids must be a list of integers.")
        
        if not TestExecution.STATUS_CHOICES: # Accessing STATUS_CHOICES requires a model instance or class
            raise ValueError("Invalid status provided.")
        valid_statuses = [choice[0] for choice in TestExecution.STATUS_CHOICES]
        if status not in valid_statuses:
            raise ValueError(f"Invalid status '{status}'. Must be one of {', '.join(valid_statuses)}.")

        test_cases = TestCase.objects.filter(id__in=test_case_ids)
        if test_cases.count() != len(test_case_ids):
            logger.warning("Some test_case_ids provided for bulk execution were not found.")

        created_count = 0
        updated_count = 0

        for tc in test_cases:
            execution, created = TestExecution.objects.update_or_create(
                testcase=tc,
                test_run=test_run, # Assuming test_run is provided or can be None for a general execution
                defaults={
                    'status': status,
                    'executor': executor_user,
                    'execution_date': timezone.now(),
                    'comments': comments,
                    # 'execution_time_minutes': None, # Can be set if timer is involved
                    # 'notes': ''
                }
            )
            if created:
                created_count += 1
            else:
                updated_count += 1
        
        message = f"Bulk execution completed: {created_count} new executions, {updated_count} updated."
        logger.info(message)
        return {'success': True, 'message': message}

    except Exception as e:
        logger.error(f"Error during bulk test case execution: {e}")
        return {'success': False, 'message': f"An error occurred: {e}"}

def execution_time_analysis(executions_queryset, period='day'):
    """
    Analyzes execution time trends over a specified period.
    Returns a dictionary suitable for charting.
    """
    if not executions_queryset.exists():
        return {'labels': [], 'avg_times': [], 'total_times': []}

    try:
        if period == 'day':
            date_col = 'execution_date__date'
        elif period == 'week':
            date_col = 'execution_date__week'
        elif period == 'month':
            date_col = 'execution_date__month'
        else:
            raise ValueError("Period must be 'day', 'week', or 'month'.")

        # Annotate with the period and then aggregate
        time_data = executions_queryset.filter(execution_date__isnull=False, execution_time_minutes__isnull=False) \
                                      .annotate(period_label=F(date_col)) \
                                      .values('period_label') \
                                      .annotate(
                                          avg_time=Avg('execution_time_minutes'),
                                          total_time=Sum('execution_time_minutes')
                                      ) \
                                      .order_by('period_label')

        labels = []
        avg_times = []
        total_times = []

        for entry in time_data:
            labels.append(entry['period_label'])
            avg_times.append(entry['avg_time'] if entry['avg_time'] is not None else 0)
            total_times.append(entry['total_time'] if entry['total_time'] is not None else 0)
        
        logger.info(f"Successfully performed execution time analysis for period: {period}.")
        return {'labels': labels, 'avg_times': avg_times, 'total_times': total_times}

    except Exception as e:
        logger.error(f"Error during execution time analysis: {e}")
        return {'labels': [], 'avg_times': [], 'total_times': [], 'error': str(e)}